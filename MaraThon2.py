import streamlit as st
import pandas as pd
from datetime import datetime, timedelta, date
import json
import os
import requests
from ics import Calendar
import io
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import random
import math

# --- REPORTLAB PRE PDF + UNICODE ---
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import urllib.request

# --- KONFIGURÁCIA ---
CONFIG_FILE = 'hospital_config.json'
HISTORY_FILE = 'room_history.json'
PRIVATE_CALENDAR_URL = "https://calendar.google.com/calendar/ical/fntnonk%40gmail.com/private-e8ce4e0639a626387fff827edd26b87f/basic.ics"
GIST_FILENAME_CONFIG = "hospital_config_v26.json"
GIST_FILENAME_HISTORY = "room_history_v26.json"
REQUEST_TIMEOUT_SECONDS = 15

ROOMS_LIST = [
    (1, 3), (2, 3), (3, 3), (4, 3), (5, 3),
    (7, 1), (8, 3), (9, 3), (10, 1), (11, 1),
    (12, 2), (13, 2), (14, 2), (15, 2), (16, 2), (17, 2),
    (18, 3), (19, 3)
]

# --- REGISTER UNICODE FONT PRE PDF ---
def setup_pdf_fonts():
    font_name = "DejaVuSans"
    if font_name in pdfmetrics.getRegisteredFontNames():
        return font_name

    font_path_candidates = [
        "/tmp/DejaVuSans.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/local/share/fonts/DejaVuSans.ttf",
        "/Library/Fonts/Arial Unicode.ttf",
        "/System/Library/Fonts/Supplemental/Arial Unicode.ttf",
        "/System/Library/Fonts/Supplemental/Arial.ttf",
    ]

    for path in font_path_candidates:
        if not os.path.exists(path):
            continue
        try:
            pdfmetrics.registerFont(TTFont(font_name, path))
            return font_name
        except:
            pass

    download_target = "/tmp/DejaVuSans.ttf"
    download_urls = [
        "https://raw.githubusercontent.com/dejavu-fonts/dejavu-fonts/master/ttf/DejaVuSans.ttf",
        "https://raw.githubusercontent.com/dejavu-fonts/dejavu-fonts/main/ttf/DejaVuSans.ttf",
    ]
    for font_url in download_urls:
        try:
            with urllib.request.urlopen(font_url, timeout=REQUEST_TIMEOUT_SECONDS) as resp:
                content = resp.read()
            with open(download_target, "wb") as f:
                f.write(content)
            pdfmetrics.registerFont(TTFont(font_name, download_target))
            return font_name
        except:
            pass
    return "Helvetica"

# --- GIST ULOŽISKO ---
def get_gist_id(filename):
    if "github" not in st.secrets: return None
    try:
        token = st.secrets["github"]["token"]
        headers = {"Authorization": f"token {token}"}
        resp = requests.get("https://api.github.com/gists", headers=headers, timeout=REQUEST_TIMEOUT_SECONDS)
        resp.raise_for_status()
        for gist in resp.json():
            if filename in gist['files']: return gist['id']
    except: pass
    return None

def load_data_from_gist(filename):
    if "github" not in st.secrets: return None
    gist_id = get_gist_id(filename)
    if not gist_id: return None
    try:
        token = st.secrets["github"]["token"]
        headers = {"Authorization": f"token {token}"}
        resp = requests.get(f"https://api.github.com/gists/{gist_id}", headers=headers, timeout=REQUEST_TIMEOUT_SECONDS)
        resp.raise_for_status()
        content = resp.json()['files'][filename]['content']
        return json.loads(content)
    except: return None

def save_data_to_gist(filename, data):
    if "github" not in st.secrets: return
    try:
        token = st.secrets["github"]["token"]
        headers = {"Authorization": f"token {token}"}
        gist_id = get_gist_id(filename)
        payload = {
            "description": f"Storage for {filename}",
            "public": False,
            "files": { filename: {"content": json.dumps(data, ensure_ascii=False, indent=2)} }
        }
        if gist_id: requests.patch(f"https://api.github.com/gists/{gist_id}", json=payload, headers=headers, timeout=REQUEST_TIMEOUT_SECONDS)
        else: requests.post("https://api.github.com/gists", json=payload, headers=headers, timeout=REQUEST_TIMEOUT_SECONDS)
    except: pass

def _load_data(gist_filename, local_filename, default_factory):
    data = load_data_from_gist(gist_filename)
    if data is not None: return data
    if os.path.exists(local_filename):
        try:
            with open(local_filename, 'r', encoding='utf-8') as f: return json.load(f)
        except: pass
    return default_factory()

def load_config():
    config = _load_data(GIST_FILENAME_CONFIG, CONFIG_FILE, get_default_config)
    config, changed = migrate_homolova_to_vidulin(config)
    if 'closures' not in config:
        config['closures'] = {}
        changed = True
    
    # Pridanie sekcie pre absencie, ak chyba
    if 'email_settings_absences' not in config:
        config['email_settings_absences'] = { 
            "default_to": "", 
            "default_subject": "Prehľad neprítomností", 
            "default_body": "Dobrý deň,\nv prílohe posielam prehľad neprítomností." 
        }
        changed = True

    for doctor_data in config.get('lekari', {}).values():
        if 'priority' not in doctor_data:
            doctor_data['priority'] = 100
            changed = True
        if 'short_term_active' not in doctor_data:
            doctor_data['short_term_active'] = (not doctor_data.get('active', True)) and bool(doctor_data.get('extra_dni', []))
            changed = True
        if 'extra_dni' not in doctor_data:
            doctor_data['extra_dni'] = []
            changed = True
        
    if changed: save_config(config)
    return config

def load_history(): return _load_data(GIST_FILENAME_HISTORY, HISTORY_FILE, lambda: {})

def save_config(config):
    try: 
        with open(CONFIG_FILE, 'w', encoding='utf-8') as f: json.dump(config, f, ensure_ascii=False, indent=2)
    except: pass
    save_data_to_gist(GIST_FILENAME_CONFIG, config)

def save_history(history):
    try: 
        with open(HISTORY_FILE, 'w', encoding='utf-8') as f: json.dump(history, f, ensure_ascii=False, indent=2)
    except: pass
    save_data_to_gist(GIST_FILENAME_HISTORY, history)

def get_default_config():
    return {
        "total_beds": 42,
        "closures": {}, 
        "email_settings": { "default_to": "", "default_subject": "Rozpis služieb", "default_body": "Dobrý deň,\nv prílohe rozpis." },
        "email_settings_absences": { "default_to": "", "default_subject": "Prehľad neprítomností", "default_body": "Dobrý deň,\nv prílohe posielam prehľad neprítomností." },
        "ambulancie": {
            "Konziliarna": { "dni": ["Pondelok", "Utorok", "Streda", "Stvrtok", "Piatok"], "priority": ["Kohutekova", "Kohutek", "Bystricky", "Zavrelova"] },
            "Velka dispenzarna": { "dni": ["Pondelok", "Utorok", "Streda", "Stvrtok", "Piatok"], "priority": ["Bocak", "Stratena", "Vidulin", "Kurisova", "Blahova", "Hrabosova", "Miklatkova", "Martinka"] },
            "Mala dispenzarna": { "dni": ["Pondelok", "Piatok"], "priority": ["Spanik", "Stratena", "Vidulin", "Kurisova", "Blahova", "Hrabosova", "Miklatkova"] },
            "Radio 2A": { "dni": ["Pondelok", "Utorok", "Streda", "Stvrtok", "Piatok"], "priority": ["Zavrelova", "Kohutek", "Kurisova", "Miklatkova", "Bystricky"], "check_presence": ["Zavrelova", "Martinka"] },
            "Radio 2B": { "dni": ["Pondelok", "Utorok", "Streda", "Stvrtok", "Piatok"], "priority": ["Martinka"], "conditional_owner": "Martinka" },
            "Chemo 8A": { "dni": ["Pondelok", "Utorok", "Streda", "Stvrtok", "Piatok"], "priority": ["Hatalova", "Kohutek", "Stratena", "Bystricky"] },
            "Chemo 8B": { "dni": ["Pondelok", "Utorok", "Streda", "Stvrtok", "Piatok"], "priority": ["Riedlova", "Kohutek", "Stratena", "Bystricky", "Vidulin", "Blahova"] },
            "Chemo 8C": { "dni": ["Utorok", "Streda", "Stvrtok"], "priority": ["Stratena", "Kohutek", "Bystricky", "Vidulin", "Blahova"] },
            "Wolf": { "dni": ["Pondelok", "Utorok", "Streda", "Stvrtok", "Piatok"], "priority": ["Spanik", "Miklatkova", "Kurisova", "Kohutek"] }
        },
        "lekari": {
            "Bystricky": { "moze": ["Konziliarna", "Velka dispenzarna", "Mala dispenzarna", "Radio 2A", "Chemo 8A", "Chemo 8B", "Chemo 8C", "Wolf"], "active": True },
            "Kohutek": { "moze": ["Oddelenie", "Konziliarna", "Velka dispenzarna", "Mala dispenzarna", "Radio 2A", "Chemo 8A", "Chemo 8B", "Chemo 8C", "Wolf"], "pevne_dni": {"Pondelok": "Chemo 8B", "Utorok": "Chemo 8B"}, "active": True },
            "Kohutekova": { "moze": ["Konziliarna"], "pevne_dni": {"Pondelok": "Konziliarna", "Utorok": "Konziliarna", "Streda": "Konziliarna", "Stvrtok": "Konziliarna"}, "nepracuje": ["Piatok"], "active": True },
            "Riedlova": { "moze": ["Chemo 8B"], "pevne_dni": {"Streda": "Chemo 8B", "Stvrtok": "Chemo 8B"}, "nepracuje": ["Pondelok", "Utorok"], "active": True },
            "Zavrelova": { "moze": ["Radio 2A", "Konziliarna"], "pevne_dni": {"Pondelok": "Radio 2A", "Utorok": "Radio 2A", "Streda": "Radio 2A", "Stvrtok": "Radio 2A", "Piatok": "Radio 2A"}, "active": True },
            "Martinka": { "moze": ["Radio 2B", "Oddelenie", "Velka dispenzarna"], "pevne_dni": {"Pondelok": "Radio 2B", "Utorok": "Radio 2B", "Streda": "Radio 2B", "Stvrtok": "Radio 2B", "Piatok": "Radio 2B"}, "active": True },
            "Hatalova": { "moze": ["Chemo 8A"], "pevne_dni": {"Pondelok": "Chemo 8A", "Utorok": "Chemo 8A", "Streda": "Chemo 8A", "Stvrtok": "Chemo 8A", "Piatok": "Chemo 8A"}, "active": True },
            "Stratena": { "moze": ["Oddelenie", "Velka dispenzarna", "Mala dispenzarna", "Chemo 8A", "Chemo 8B", "Chemo 8C"], "pevne_dni": {"Utorok": "Chemo 8C", "Streda": "Chemo 8C", "Stvrtok": "Chemo 8C"}, "active": True },
            "Vidulin": { "moze": ["Oddelenie", "Velka dispenzarna", "Mala dispenzarna", "Chemo 8A", "Chemo 8B", "Chemo 8C"], "active": True },
            "Miklatkova": { "moze": ["Oddelenie", "Wolf"], "active": True },
            "Kurisova": { "moze": ["Oddelenie", "Velka dispenzarna", "Mala dispenzarna", "Radio 2A", "Wolf"], "special": "veduca", "active": True },
            "Blahova": { "moze": ["Oddelenie", "Velka dispenzarna", "Mala dispenzarna", "Chemo 8B", "Chemo 8C"], "active": False },
            "Hrabosova": { "moze": ["Oddelenie", "Velka dispenzarna", "Mala dispenzarna"], "active": False, "short_term_active": True, "extra_dni": [] },
            "Bocak": { "moze": ["Velka dispenzarna"], "pevne_dni": {"Pondelok": "Velka dispenzarna", "Utorok": "Velka dispenzarna", "Streda": "Velka dispenzarna", "Stvrtok": "Velka dispenzarna", "Piatok": "Velka dispenzarna"}, "active": True },
            "Spanik": { "moze": ["Wolf", "Mala dispenzarna"], "pevne_dni": {"Pondelok": "Mala dispenzarna", "Utorok": "Wolf", "Streda": "Wolf", "Stvrtok": "Wolf", "Piatok": "Mala dispenzarna"}, "active": True },
            "Kacurova": { "moze": ["Oddelenie"], "active": True },
            "Hunakova": { "moze": ["Oddelenie"], "active": True }
        }
    }

def migrate_homolova_to_vidulin(config):
    changed = False
    if "Homolova" in config["lekari"]:
        config["lekari"]["Vidulin"] = config["lekari"].pop("Homolova")
        changed = True
    for amb_name, amb_data in config["ambulancie"].items():
        if isinstance(amb_data["priority"], list):
            if "Homolova" in amb_data["priority"]:
                amb_data["priority"] = ["Vidulin" if x == "Homolova" else x for x in amb_data["priority"]]
                changed = True
        elif isinstance(amb_data["priority"], dict):
            for day_key, day_list in amb_data["priority"].items():
                if "Homolova" in day_list:
                    amb_data["priority"][day_key] = ["Vidulin" if x == "Homolova" else x for x in day_list]
                    changed = True
    return config, changed

def is_doctor_active_on_date(doc_props, date_key):
    if doc_props.get('active', True):
        return True
    return doc_props.get('short_term_active', False) and date_key in doc_props.get('extra_dni', [])

def _replace_doctor_references(config, old_name, new_name=None, remove=False):
    for amb_data in config.get("ambulancie", {}).values():
        prio = amb_data.get("priority")
        if isinstance(prio, list):
            updated = []
            for doc_name in prio:
                if doc_name == old_name:
                    if remove:
                        continue
                    updated.append(new_name)
                else:
                    updated.append(doc_name)
            amb_data["priority"] = updated
        elif isinstance(prio, dict):
            for day_key, day_list in prio.items():
                updated = []
                for doc_name in day_list:
                    if doc_name == old_name:
                        if remove:
                            continue
                        updated.append(new_name)
                    else:
                        updated.append(doc_name)
                prio[day_key] = updated

        if isinstance(amb_data.get("check_presence"), list):
            if remove:
                amb_data["check_presence"] = [x for x in amb_data["check_presence"] if x != old_name]
            else:
                amb_data["check_presence"] = [new_name if x == old_name else x for x in amb_data["check_presence"]]

        if amb_data.get("conditional_owner") == old_name:
            amb_data["conditional_owner"] = None if remove else new_name

def _rename_doctor_in_history(history, old_name, new_name):
    for day_key, day_map in history.items():
        if old_name in day_map:
            day_map[new_name] = day_map.pop(old_name)

def _rename_doctor_in_manual_core(manual_core, old_name, new_name):
    for date_key, day_map in manual_core.items():
        if old_name in day_map:
            day_map[new_name] = day_map.pop(old_name)

def rename_doctor_everywhere(config, history, manual_core, old_name, new_name):
    if old_name not in config.get("lekari", {}):
        return False, "Lekár neexistuje."
    new_name = (new_name or "").strip()
    if not new_name:
        return False, "Nové meno je prázdne."
    if new_name != old_name and new_name in config["lekari"]:
        return False, "Lekár s týmto menom už existuje."
    if new_name == old_name:
        return True, ""

    config["lekari"][new_name] = config["lekari"].pop(old_name)
    _replace_doctor_references(config, old_name, new_name=new_name, remove=False)
    _rename_doctor_in_history(history, old_name, new_name)
    _rename_doctor_in_manual_core(manual_core, old_name, new_name)
    return True, ""

def remove_doctor_everywhere(config, history, manual_core, doctor_name):
    if doctor_name not in config.get("lekari", {}):
        return False, "Lekár neexistuje."
    config["lekari"].pop(doctor_name, None)
    _replace_doctor_references(config, doctor_name, remove=True)
    for day_map in history.values():
        day_map.pop(doctor_name, None)
    for day_map in manual_core.values():
        day_map.pop(doctor_name, None)
    return True, ""

def distribute_rooms(doctors_list, wolf_doc_name, previous_assignments=None, manual_preferences=None, doctor_priorities=None):
    if not doctors_list: return {}, {}
    if manual_preferences is None: manual_preferences = {}
    if previous_assignments is None: previous_assignments = {}
    if doctor_priorities is None: doctor_priorities = {}

    total_beds = sum(r[1] for r in ROOMS_LIST)
    min_docs_for_limit = math.ceil(total_beds / 15)
    protected = ["Miklatkova", "Kohutek", "Kurisova"]
    non_protected = [d for d in doctors_list if d not in protected]
    use_for_rooms = list(non_protected)

    if len(non_protected) < min_docs_for_limit:
        for fallback_doc in protected:
            if fallback_doc in doctors_list and fallback_doc not in use_for_rooms:
                use_for_rooms.append(fallback_doc)
            if len(use_for_rooms) >= min_docs_for_limit:
                break
    elif len(non_protected) >= 3:
        use_for_rooms = list(non_protected)

    no_room_docs = [d for d in doctors_list if d not in use_for_rooms]
    if not use_for_rooms:
        use_for_rooms = list(doctors_list)
        no_room_docs = []

    assignment = {d: [] for d in doctors_list}
    current_beds = {d: 0 for d in doctors_list}
    available_rooms = sorted(ROOMS_LIST, key=lambda x: x[0]) 

    # --- 1. TARGET CALCULATION (equal load, respect 15 where possible) ---
    targets = {d: 0 for d in doctors_list}
    base = total_beds // len(use_for_rooms)
    rem = total_beds % len(use_for_rooms)
    ordered_for_target = sorted(use_for_rooms, key=lambda d: (doctor_priorities.get(d, 100), d))
    for i, doc in enumerate(ordered_for_target):
        targets[doc] = base + (1 if i < rem else 0)

    # --- 2. PREFERENCIE (MANUAL) ---
    for doc, nums in manual_preferences.items():
        if doc not in use_for_rooms: continue
        my_target = targets.get(doc, 15)
        for num in nums:
            r_obj = next((r for r in available_rooms if r[0] == num), None)
            if not r_obj: continue
            if current_beds[doc] + r_obj[1] <= my_target:
                assignment[doc].append(r_obj)
                current_beds[doc] += r_obj[1]
                available_rooms.remove(r_obj)

    # --- 3. KONTINUITA (PREVIOUS) ---
    if previous_assignments:
        for doc in use_for_rooms:
            if doc in previous_assignments:
                my_prev = []
                for r_id in previous_assignments[doc]:
                    found = next((r for r in available_rooms if r[0] == r_id), None)
                    if found: my_prev.append(found)
                my_target = targets.get(doc, 15)
                random.shuffle(my_prev)
                for r_obj in my_prev:
                    if current_beds[doc] + r_obj[1] <= my_target:
                        assignment[doc].append(r_obj)
                        current_beds[doc] += r_obj[1]
                        available_rooms.remove(r_obj)

    # --- 4. DOROVNÁVANIE ---
    while available_rooms:
        candidates = [d for d in use_for_rooms if current_beds[d] < targets.get(d, 15)]
        if not candidates: candidates = list(use_for_rooms)
        candidates.sort(key=lambda d: (current_beds[d], doctor_priorities.get(d, 100), d))
        receiver = candidates[0]
        best_room = available_rooms[0]
        if assignment[receiver]:
            avgs = sum(r[0] for r in assignment[receiver]) / len(assignment[receiver])
            best_room = min(available_rooms, key=lambda r: abs(r[0] - avgs))
        assignment[receiver].append(best_room)
        current_beds[receiver] += best_room[1]
        available_rooms.remove(best_room)

    result_text, result_raw = {}, {}
    for doc in doctors_list:
        rooms = sorted(assignment[doc], key=lambda x: x[0])
        result_raw[doc] = [r[0] for r in rooms]
        r_str = ", ".join([str(r[0]) for r in rooms])

        if not rooms:
             if doc in no_room_docs and doc in ["Miklatkova", "Kohutek", "Kurisova"]:
                 result_text[doc] = "RT oddelenie"
             elif doc == wolf_doc_name:
                 result_text[doc] = "Wolf (0L)"
             else:
                 result_text[doc] = ""
        else:
             result_text[doc] = f"{r_str} + Wolf" if doc == wolf_doc_name else r_str
             
    return result_text, result_raw

def get_ical_events(start_date, end_date):
    try:
        response = requests.get(PRIVATE_CALENDAR_URL, timeout=REQUEST_TIMEOUT_SECONDS)
        response.raise_for_status()
        c = Calendar(response.text)
        absences = {}
        for event in c.events:
            ev_start, ev_end = event.begin.date(), event.end.date()
            if ev_end < start_date.date() or ev_start > end_date.date(): continue
            raw = event.name.strip()
            name, typ = raw, "Dovolenka"
            if raw.upper().endswith('PN'): typ, name = "PN", raw[:-2].rstrip(' -')
            elif raw.upper().endswith('VZ'): typ, name = "Vzdelávanie", raw[:-2].rstrip(' -')
            elif raw.upper().endswith('S') and not raw.upper().endswith('OS'): typ, name = "Stáž", raw[:-1].rstrip(' -')
            elif '-' in raw and typ == "Dovolenka":
                parts = raw.split('-')
                name = parts[0].strip()
            curr, limit = max(ev_start, start_date.date()), min(ev_end, end_date.date())
            while curr < limit:
                absences.setdefault(curr.strftime('%Y-%m-%d'), {})[name] = typ
                curr += timedelta(days=1)
        return absences
    except: return {}

def build_absence_table(absences, start_d):
    # Oprava typu vstupu
    if isinstance(start_d, datetime):
        current_start_date = start_d.date()
    else:
        current_start_date = start_d
        
    date_range_start = current_start_date
    date_range_end = current_start_date + timedelta(days=7)
    
    raw_entries = []
    # Prechadzame vsetky kluce v absences
    for date_str in sorted(absences.keys()):
        d_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
        if date_range_start <= d_obj <= date_range_end:
            for person, reason in absences[date_str].items():
                raw_entries.append({
                    "date": d_obj,
                    "person": person,
                    "reason": reason
                })
    
    if not raw_entries:
        return pd.DataFrame(columns=["Od - Do", "Lekár", "Dôvod"])

    raw_entries.sort(key=lambda x: (x['person'], x['reason'], x['date']))
    
    grouped_rows = []
    if raw_entries:
        current_person = raw_entries[0]['person']
        current_reason = raw_entries[0]['reason']
        start_date = raw_entries[0]['date']
        last_date = raw_entries[0]['date']

        for entry in raw_entries[1:]:
            is_consecutive = (entry['date'] - last_date).days == 1
            is_same_group = (entry['person'] == current_person and entry['reason'] == current_reason)
            
            if is_same_group and is_consecutive:
                last_date = entry['date']
            else:
                date_str = f"{start_date.strftime('%d.%m.')} - {last_date.strftime('%d.%m.%Y')}" if start_date != last_date else f"{start_date.strftime('%d.%m.%Y')}"
                grouped_rows.append({
                    "Od - Do": date_str,
                    "Lekár": current_person,
                    "Dôvod": current_reason
                })
                current_person = entry['person']
                current_reason = entry['reason']
                start_date = entry['date']
                last_date = entry['date']

        date_str = f"{start_date.strftime('%d.%m.')} - {last_date.strftime('%d.%m.%Y')}" if start_date != last_date else f"{start_date.strftime('%d.%m.%Y')}"
        grouped_rows.append({
            "Od - Do": date_str,
            "Lekár": current_person,
            "Dôvod": current_reason
        })

    return pd.DataFrame(grouped_rows)

def generate_data_structure(config, absences, start_date, save_hist=True):
    days_map = {0: "Pondelok", 1: "Utorok", 2: "Streda", 3: "Stvrtok", 4: "Piatok"}
    weekday = start_date.weekday()
    thursday = start_date + timedelta(days=(3 - weekday) % 7)
    dates, data_grid = [], {}
    all_doctors, doctors_info = [], {}
    week_dates_str = []
    for i in range(7):
        d = thursday + timedelta(days=i)
        if d.weekday() < 5: week_dates_str.append(d.strftime('%Y-%m-%d'))

    for d_name, props in config['lekari'].items():
        short_term_hits = [ed for ed in props.get('extra_dni', []) if ed in week_dates_str]
        if props.get('active', True) or (props.get('short_term_active', False) and short_term_hits):
            all_doctors.append(d_name)
            if not props.get('active', True) and props.get('short_term_active', False):
                readable = [datetime.strptime(ed, '%Y-%m-%d').strftime('%d.%m.') for ed in short_term_hits]
                doctors_info[d_name] = f"⚠️ len {', '.join(readable)}"

    all_doctors.sort()
    doctor_priorities = {d: config['lekari'].get(d, {}).get('priority', 100) for d in all_doctors}
    history = load_history()
    last_day_assignments = history.get((thursday - timedelta(days=1)).strftime('%Y-%m-%d'), {})
    manual_all = st.session_state.get("manual_core", {})
    closures = config.get('closures', {})
    
    dates_raw = []

    for i in range(7):
        curr_date = thursday + timedelta(days=i)
        day_name = days_map.get(curr_date.weekday())
        if not day_name: continue
        date_str = curr_date.strftime('%d.%m.%Y')
        date_key = curr_date.strftime('%Y-%m-%d')
        dates.append(date_str)
        dates_raw.append(date_key)
        day_absences = absences.get(date_key, {})
        closed_today = closures.get(date_key, [])
        data_grid[date_str] = {}
        
        available = [
            d for d in all_doctors
            if is_doctor_active_on_date(config['lekari'][d], date_key)
            and d not in day_absences
            and day_name not in config['lekari'][d].get('nepracuje', [])
        ]
        assigned_amb = {}
        
        for doc in list(available):
            if fixed := config['lekari'][doc].get('pevne_dni', {}).get(day_name):
                for t in [t.strip() for t in fixed.split(',')]:
                    if t in closed_today: assigned_amb[t] = "ZATVORENÉ"
                    else: assigned_amb[t] = doc
                available.remove(doc)
        
        ambs_to_process = ["Radio 2A", "Radio 2B", "Chemo 8B", "Chemo 8A", "Chemo 8C", "Wolf", "Konziliarna", "Velka dispenzarna", "Mala dispenzarna"]
        amb_scarcity = []
        for amb_name in ambs_to_process:
            if amb_name in assigned_amb or amb_name in closed_today: 
                if amb_name in closed_today: assigned_amb[amb_name] = "ZATVORENÉ"
                continue
            if day_name not in config['ambulancie'][amb_name]['dni']:
                assigned_amb[amb_name] = "---"
                continue
            if amb_name == "Radio 2B" and "Martinka" not in available:
                assigned_amb[amb_name] = "ZATVORENÉ"
                continue
            prio = config['ambulancie'][amb_name]['priority']
            if isinstance(prio, dict): prio = prio.get(str(curr_date.weekday()), prio.get('default', []))
            cands = [d for d in prio if d in available and amb_name in config['lekari'][d].get('moze', [])]
            amb_scarcity.append({"name": amb_name, "candidates": cands, "count": len(cands), "idx": ambs_to_process.index(amb_name)})
        
        amb_scarcity.sort(key=lambda x: (x['count'], x['idx']))
        for item in amb_scarcity:
            amb = item['name']
            cands = [c for c in item['candidates'] if c in available]
            if amb == "Wolf" and "Spanik" in all_doctors and "Spanik" not in day_absences and assigned_amb.get("Mala dispenzarna") == "Spanik":
                 assigned_amb["Wolf"] = "Spanik"
                 continue
            if not cands:
                assigned_amb[amb] = "NEOBSADENÉ"
                continue
            cands = sorted(
                cands,
                key=lambda d: (
                    1 if d == "Bystricky" else 0,
                    doctor_priorities.get(d, 100),
                    item['candidates'].index(d)
                )
            )
            chosen = cands[0]
            assigned_amb[amb] = chosen
            available.remove(chosen)

        # Na 2A/2B staci obsadit aspon jednu RT ambulanciu.
        non_active_values = {"NEOBSADENÉ", "ZATVORENÉ", "---", ""}
        r2a = assigned_amb.get("Radio 2A", "")
        r2b = assigned_amb.get("Radio 2B", "")
        if r2a not in non_active_values and r2b == "NEOBSADENÉ":
            assigned_amb["Radio 2B"] = "---"
        if r2b not in non_active_values and r2a == "NEOBSADENÉ":
            assigned_amb["Radio 2A"] = "---"

        for k, v in assigned_amb.items(): data_grid[date_str][k] = v
        
        wolf_doc = assigned_amb.get("Wolf")
        if "ODDELENIE (Celé)" in closed_today:
            room_text_map, room_raw_map = {}, {}
            for d in all_doctors: room_text_map[d] = "ZATVORENÉ"
        else:
            ward_cands = [d for d in available if "Oddelenie" in config['lekari'][d].get('moze', [])]
            if wolf_doc and wolf_doc not in ward_cands and "Oddelenie" in config['lekari'].get(wolf_doc, {}).get('moze', []):
                ward_cands.append(wolf_doc)
            
            daily_pref = manual_all.get(date_key, {})
            if not daily_pref:
                start_key = start_date.strftime('%Y-%m-%d')
                daily_pref = manual_all.get(start_key, {})

            room_text_map, room_raw_map = distribute_rooms(
                ward_cands,
                wolf_doc,
                last_day_assignments,
                daily_pref,
                doctor_priorities=doctor_priorities
            )
            last_day_assignments = room_raw_map
            if save_hist: history[date_key] = room_raw_map
        
        for doc in all_doctors:
            if not is_doctor_active_on_date(config['lekari'][doc], date_key):
                data_grid[date_str][doc] = ""
                continue
            if doc in day_absences: data_grid[date_str][doc] = day_absences[doc]
            elif doc in room_text_map: data_grid[date_str][doc] = room_text_map[doc]
            else:
                my = [a for a, d in assigned_amb.items() if d == doc]
                data_grid[date_str][doc] = " + ".join(my) if my else ""
                
    if save_hist:
        save_history(history)
    return dates, data_grid, all_doctors, doctors_info, dates_raw

def scan_future_problems(config, weeks_ahead=12):
    problems = []
    start = datetime.now()
    end = start + timedelta(weeks=weeks_ahead)
    absences = get_ical_events(start, end)
    closures = config.get('closures', {})
    current = start
    while current <= end:
        dates, grid, docs, info, _ = generate_data_structure(config, absences, current, save_hist=False)
        for date_str in dates:
            date_obj = datetime.strptime(date_str, '%d.%m.%Y')
            date_key = date_obj.strftime('%Y-%m-%d')
            closed_today = closures.get(date_key, [])
            for amb_name in ["Konziliarna", "Velka dispenzarna", "Mala dispenzarna", "Radio 2A", "Radio 2B", "Chemo 8A", "Chemo 8B", "Chemo 8C", "Wolf"]:
                val = grid[date_str].get(amb_name, "")
                if val in ["NEOBSADENÉ", "???", ""] and amb_name not in closed_today and "ODDELENIE (Celé)" not in closed_today:
                     problems.append({"Dátum": date_str, "Pracovisko": amb_name})
        current += timedelta(weeks=1)
    return pd.DataFrame(problems) if problems else None

def create_display_df(dates, data_grid, all_doctors, doctors_info, motto, config):
    rows = []
    ward_doctors = [d for d in all_doctors if "Oddelenie" in config['lekari'][d].get('moze', [])]
    display_map = { "Radio 2A": "Radio 2A", "Konziliarna": "Konziliárna amb.", "Velka dispenzarna": "veľký dispenzár", "Mala dispenzarna": "malý dispenzár" }
    
    rows.append(["Oddelenie"] + dates)
    for doc in ward_doctors:
        vals = []
        for date in dates:
            val = data_grid[date].get(doc, "")
            for old, new in display_map.items(): val = val.replace(old, new)
            vals.append(val)
        label = f"Dr {doc}" + (f" {doctors_info[doc]}" if doc in doctors_info else "")
        rows.append([label] + vals)
    rows.append([motto or "Motto"] + [""] * len(dates))
    sections = [("Konziliárna amb", ["Konziliarna"]), ("RT ambulancie", ["Radio 2A", "Radio 2B"]), ("Chemo amb", ["Chemo 8A", "Chemo 8B", "Chemo 8C"]), ("Disp. Ambulancia", ["Velka dispenzarna", "Mala dispenzarna"]), ("RTG Terapia", ["Wolf"])]
    for title, ambs in sections:
        rows.append([title] + dates)
        for amb in ambs:
            vals = [data_grid[d].get(amb, "").replace("---", "").replace("NEOBSADENÉ", "???") for d in dates]
            rows.append([display_map.get(amb, amb)] + vals)
        rows.append([""] * (len(dates) + 1))
    return pd.DataFrame(rows)

def create_excel_report(df):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, header=False, sheet_name="Rozpis")
        ws = writer.sheets['Rozpis']
        bold, center, thin = Font(bold=True), Alignment(horizontal="center", vertical="center", wrap_text=True), Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        ws.cell(1, 1, f"Rozpis prác Onkologická klinika {df.columns[1]} - {df.columns[-1]}").font = bold
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
        ws['A1'].alignment = center
        for r, row in enumerate(df.iterrows(), 2):
            is_header = row[1][0] in ["Oddelenie", "Konziliárna amb", "RT ambulancie", "Chemo amb", "Disp. Ambulancia", "RTG Terapia"]
            is_motto = (row[1][0] == st.session_state.get('motto', 'Motto'))
            for c, val in enumerate(row[1], 1):
                cell = ws.cell(r, c, val)
                cell.border = thin
                cell.alignment = center
                if is_header or (c==1 and not is_motto): cell.font = bold
                if is_motto:
                    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(df.columns))
                    cell.font, cell.fill = Font(bold=True, italic=True), PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
                    ws.row_dimensions[r].height = 25
                    break
        ws.column_dimensions['A'].width = 25
        for i in range(2, len(df.columns) + 1): ws.column_dimensions[get_column_letter(i)].width = 18
    return output.getvalue()

def create_pdf_report(df, motto, title_prefix="Rozpis prác"):
    buffer = io.BytesIO()
    font_name = setup_pdf_fonts()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=10, leftMargin=10, topMargin=10, bottomMargin=10)
    styles = getSampleStyleSheet()
    styles['Title'].fontName = font_name
    styles['Normal'].fontName = font_name
    cell_style = ParagraphStyle('C', parent=styles['Normal'], fontName=font_name, fontSize=7, leading=8, alignment=1)
    
    # Pre absencie pouzivame inu strukturu, pre rozpis inu
    # Ak je to absencna tabulka, je jednoduchsia
    if "Od - Do" in df.columns:
        # Absencna tabulka ma menej stlpcov
        data = [[Paragraph(str(c), ParagraphStyle('H', parent=styles['Normal'], fontName=font_name, fontSize=8, alignment=1)) for c in df.columns]]
        for _, row in df.iterrows():
            row_data = []
            for val in row.values:
                row_data.append(Paragraph(str(val), cell_style))
            data.append(row_data)
        
        # Sirky stlpcov pre absencie
        col_widths = [150, 200, 200]
        if len(df.columns) != 3: col_widths = None # fallback
        
        t = Table(data, colWidths=col_widths)
        style = TableStyle([
            ('GRID', (0,0), (-1,-1), 0.5, colors.black),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('BACKGROUND', (0,0), (-1,0), colors.lightgrey)
        ])
        t.setStyle(style)
        
        # Titulok bez datumu, kedze ten je v riadkoch
        doc.build([Paragraph(f"{title_prefix}", styles['Title']), t])
        
    else:
        # Klasicky rozpis
        data = [[Paragraph(str(c), ParagraphStyle('H', parent=styles['Normal'], fontName=font_name, fontSize=8, alignment=1)) for c in df.columns]]
        for _, row in df.iterrows():
            row_data = []
            is_motto = (row[0] == (motto or "Motto"))
            for i, val in enumerate(row.values):
                txt = str(val) if val else ""
                if is_motto and i==0: 
                    p = Paragraph(f"<para align='center'><b><i>{txt}</i></b></para>", ParagraphStyle('M', parent=cell_style, fontSize=9, padding=6, alignment=1))
                elif is_motto: p = ""
                elif i==0: p = Paragraph(f"<b>{txt}</b>", cell_style)
                else: p = Paragraph(txt, cell_style)
                row_data.append(p)
            data.append(row_data)
        
        t = Table(data, colWidths=[130] + [135]*(len(df.columns)-1))
        style = TableStyle([('GRID', (0,0), (-1,-1), 0.5, colors.black), ('VALIGN', (0,0), (-1,-1), 'MIDDLE'), ('BACKGROUND', (0,0), (-1,0), colors.grey)])
        for i, row in enumerate(df.iterrows()):
            if row[1][0] in ["Oddelenie", "Konziliárna amb", "RT ambulancie", "Chemo amb", "Disp. Ambulancia", "RTG Terapia"]:
                style.add('BACKGROUND', (0, i+1), (-1, i+1), colors.lightgrey)
            if row[1][0] == (motto or "Motto"):
                style.add('SPAN', (0, i+1), (-1, i+1))
                style.add('BACKGROUND', (0, i+1), (-1, i+1), colors.whitesmoke)
                style.add('ALIGN', (0, i+1), (-1, i+1), 'CENTER')
        t.setStyle(style)
        
        doc.build([Paragraph(f"{title_prefix} {df.columns[1]} - {df.columns[-1]}", styles['Title']), t])
        
    buffer.seek(0)
    return buffer.getvalue()

def send_email_with_pdf(pdf_bytes, filename, to_email, subject, body):
    if "email" not in st.secrets:
        return False, "Chýba sekcia [email] v secrets."
    if not to_email or not to_email.strip():
        return False, "Nie je zadaný príjemca emailu."
    try:
        sender = st.secrets["email"]["username"]
        password = st.secrets["email"]["password"]
        msg = MIMEMultipart()
        msg['From'], msg['To'], msg['Subject'] = sender, to_email, subject
        msg.attach(MIMEText(body, 'plain'))
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(pdf_bytes)
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f'attachment; filename={filename}')
        msg.attach(part)
        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        server.login(sender, password)
        server.send_message(msg)
        server.quit()
        return True, ""
    except KeyError as e:
        return False, f"Chýba kľúč v secrets: {e}"
    except Exception as e:
        return False, str(e)

def group_closures_to_intervals(closures_dict):
    sorted_dates = sorted(closures_dict.keys())
    if not sorted_dates: return []
    intervals, curr_start, curr_end, curr_val = [], sorted_dates[0], sorted_dates[0], sorted(closures_dict[sorted_dates[0]])
    for d_str in sorted_dates[1:]:
        d, prev = datetime.strptime(d_str, '%Y-%m-%d').date(), datetime.strptime(curr_end, '%Y-%m-%d').date()
        val = sorted(closures_dict[d_str])
        if (d - prev).days == 1 and val == curr_val: curr_end = d_str
        else:
            intervals.append((curr_start, curr_end, curr_val))
            curr_start, curr_end, curr_val = d_str, d_str, val
    intervals.append((curr_start, curr_end, curr_val))
    return intervals

# --- MAIN APP ---
st.set_page_config(page_title="Rozpis FN Trenčín", layout="wide")
st.title("🏥 Rozpis prác - Onkologická klinika FN Trenčín")

if 'config' not in st.session_state: st.session_state.config = load_config()
if 'manual_core' not in st.session_state: st.session_state.manual_core = {}
if 'temp_exceptions' not in st.session_state: st.session_state.temp_exceptions = []

mode = st.sidebar.radio("Navigácia", ["🚀 Generovať rozpis", "⚙️ Nastavenia lekárov", "🏥 Nastavenia ambulancií", "📧 Nastavenia Emailu"])

if mode == "🚀 Generovať rozpis":
    c1, c2 = st.columns(2)
    st.session_state.motto = c1.text_input("📢 Motto:", placeholder="...")
    start_d = c2.date_input("Začiatok:", datetime.now())

    with st.expander("📅 Výnimky", expanded=True):
        if st.session_state.config.get('closures'):
            st.markdown("##### 💾 Aktívne výnimky:")
            for s, e, l in group_closures_to_intervals(st.session_state.config['closures']):
                c1, c2, c3 = st.columns([2, 4, 1])
                lbl = f"{datetime.strptime(s, '%Y-%m-%d').strftime('%d.%m.')} - {datetime.strptime(e, '%Y-%m-%d').strftime('%d.%m.')}" if s!=e else datetime.strptime(s, '%Y-%m-%d').strftime('%d.%m.')
                c1.text(lbl)
                c2.text(", ".join(l) if len(l)<5 else "VIACERÉ")
                if c3.button("🗑️", key=f"d_{s}"):
                    curr = datetime.strptime(s, '%Y-%m-%d')
                    while curr <= datetime.strptime(e, '%Y-%m-%d'):
                        st.session_state.config['closures'].pop(curr.strftime('%Y-%m-%d'), None)
                        curr += timedelta(days=1)
                    save_config(st.session_state.config)
                    st.rerun()

        st.markdown("---")
        c1, c2 = st.columns([1, 2])
        nr = c1.date_input("Nový rozsah:", value=[], key="n_r")
        nc = c2.multiselect("Zatvoriť:", ["ODDELENIE (Celé)"] + list(st.session_state.config['ambulancie'].keys()), key="n_c")
        if st.button("➕ Pridať"):
            if nr and nc:
                st.session_state.temp_exceptions.append(((nr[0], nr[1] if len(nr)>1 else nr[0]), nc))
                st.rerun()

        if st.session_state.temp_exceptions:
            st.write("Nové (neuložené):")
            for i, (r, l) in enumerate(st.session_state.temp_exceptions):
                st.text(f"{r[0]} - {r[1]}: {l}")
            if st.button("💾 Uložiť všetko"):
                for r, l in st.session_state.temp_exceptions:
                    c = r[0]
                    while c <= r[1]:
                        k = c.strftime('%Y-%m-%d')
                        st.session_state.config['closures'][k] = list(set(st.session_state.config['closures'].get(k, []) + l))
                        c += timedelta(days=1)
                st.session_state.temp_exceptions = []
                save_config(st.session_state.config)
                st.rerun()

    st.markdown("### Manuálne pridelenie izieb")
    manual_core_input = {}
    ward_docs = [
        d for d, p in st.session_state.config["lekari"].items()
        if "Oddelenie" in p.get("moze", [])
        and (p.get("active", True) or p.get("short_term_active", False))
    ]

    if st.button("📥 Načítať izby z minulého týždňa"):
        curr_week_start = start_d + timedelta(days=(3 - start_d.weekday()) % 7)
        prev_week_key = (curr_week_start - timedelta(days=7)).strftime('%Y-%m-%d')
        history_for_load = load_history()
        prev_rooms = history_for_load.get(prev_week_key, {})
        if prev_rooms:
            loaded_manual = {}
            for doc in ward_docs:
                rooms = prev_rooms.get(doc, [])
                if rooms:
                    room_nums = [int(x) for x in rooms if isinstance(x, int) or (isinstance(x, str) and str(x).isdigit())]
                    if room_nums:
                        loaded_manual[doc] = room_nums
                        st.session_state[f"core_{doc}"] = ", ".join(str(x) for x in room_nums)
            if loaded_manual:
                st.session_state.manual_core[start_d.strftime('%Y-%m-%d')] = loaded_manual
                st.success(f"Načítané izby z dátumu {prev_week_key}.")
            else:
                st.warning(f"V histórii ({prev_week_key}) neboli nájdené izby pre aktívnych lekárov oddelenia.")
        else:
            st.warning(f"V histórii nie je záznam pre {prev_week_key}.")

    cols = st.columns(2)
    for i, doc in enumerate(ward_docs):
        with cols[i % 2]:
            val = st.text_input(f"Dr {doc} – izby (napr. 1, 4):", key=f"core_{doc}")
            if val.strip():
                try:
                    manual_core_input[doc] = [int(p.strip()) for p in val.split(',') if p.strip().isdigit()]
                except: pass
    
    if manual_core_input:
        st.session_state.manual_core[start_d.strftime('%Y-%m-%d')] = manual_core_input

    c_btn1, c_btn2, c_btn3 = st.columns(3)
    gen_clicked = c_btn1.button("🚀 Generovať rozpis", type="primary")
    scan_clicked = c_btn2.button("🔭 Vyhliadka ďalších týždňov")
    clear_hist = c_btn3.button("🗑️ Reset histórie")
    weeks_num = st.number_input("Počet týždňov pre vyhliadku:", min_value=1, max_value=52, value=12)

    if gen_clicked:
        with st.spinner("..."):
            end_d = start_d + timedelta(days=14)
            ab = get_ical_events(datetime.combine(start_d, datetime.min.time()), datetime.combine(end_d, datetime.min.time()))
            ds, g, d, di, raw_dates = generate_data_structure(st.session_state.config, ab, start_d)
            st.session_state.dates_raw = raw_dates
            st.session_state.df_generated = create_display_df(ds, g, d, di, st.session_state.motto, st.session_state.config)
            st.session_state.df_generated.columns = ["Sekcia / Dátum"] + ds
            st.session_state.absences_df = build_absence_table(ab, start_d)
        st.success("Hotovo!")
    
    if scan_clicked:
        with st.spinner(f"Pozerám {weeks_num} týždňov dopredu..."):
            problems_df = scan_future_problems(st.session_state.config, weeks_ahead=weeks_num)
            if problems_df is not None and not problems_df.empty:
                st.subheader("🔭 Vyhliadka ďalších týždňov – problémové dni")
                st.dataframe(problems_df, use_container_width=True, hide_index=True)
            else:
                st.success("✅ V zadanom období nie sú žiadne neobsadené pracoviská.")

    if clear_hist:
        save_history({})
        st.success("História zmazaná")

    if 'df_generated' in st.session_state:
        st.markdown("---")
        
        # --- SEKCE PRE ABSENCIE ---
        if 'absences_df' in st.session_state and not st.session_state.absences_df.empty:
            with st.expander("📋 Prehľad neprítomností (Dovolenky, PN, Stáže)", expanded=True):
                st.dataframe(st.session_state.absences_df, use_container_width=True, hide_index=True)
                
                # Samostatne odosielanie pre absencie
                st.markdown("#### 📧 Odoslať zoznam absencií emailom")
                c_a1, c_a2 = st.columns([3, 1])
                email_abs = st.session_state.config.get('email_settings_absences', {})
                to_abs = c_a1.text_input("Komu (Absencie):", email_abs.get('default_to', ''), key="to_abs")
                
                if c_a2.button("Odoslať Absencie"):
                    # Generovanie PDF pre absencie
                    pdf_abs = create_pdf_report(st.session_state.absences_df, None, title_prefix="Prehľad neprítomností")
                    fn_abs = f"Nepritomnosti_{start_d.strftime('%d.%m.')}.pdf"
                    subj_abs = email_abs.get('default_subject', "Nepritomnosti")
                    body_abs = email_abs.get('default_body', "V prilohe.")
                    
                    ok, err = send_email_with_pdf(pdf_abs, fn_abs, to_abs, subj_abs, body_abs)
                    if ok:
                        st.success(f"Absencie odoslané na {to_abs}")
                    else:
                        st.error(f"Chyba pri odosielaní absencií: {err}")

        # --- SEKCE PRE HLAVNY ROZPIS ---
        st.info("✏️ Tabuľku nižšie môžete priamo editovať. Zmeny sa prejavia v exportoch.")
        
        edited_df = st.data_editor(
            st.session_state.df_generated,
            use_container_width=True,
            hide_index=True,
            key="final_editor"
        )
        
        if 'dates_raw' in st.session_state:
             if st.button("💾 Uložiť aktuálne rozdelenie izieb do histórie (kontinuita)"):
                try:
                    history = load_history()
                    cols = edited_df.columns
                    for i, date_key in enumerate(st.session_state.dates_raw):
                        col_idx = i + 1
                        if col_idx >= len(cols): break
                        
                        col_name = cols[col_idx]
                        day_map = {}
                        
                        for idx, row in edited_df.iterrows():
                            label = str(row[cols[0]])
                            if label.startswith("Dr "):
                                doc_name = label.replace("Dr ", "").split(" ")[0].strip()
                                cell_val = str(row[col_name])
                                room_part = cell_val.split('+')[0]
                                nums = []
                                for piece in room_part.split(','):
                                    piece = piece.strip()
                                    if piece.isdigit():
                                        nums.append(int(piece))
                                
                                if nums:
                                    day_map[doc_name] = nums
                        
                        if day_map:
                            history[date_key] = day_map
                    
                    save_history(history)
                    st.success("✅ Rozdelenie izieb bolo uložené. Ďalšie generovanie bude nadväzovať na tieto zmeny.")
                except Exception as e:
                    st.error(f"Chyba pri ukladaní: {e}")

        export_df = edited_df.copy()
        xlsx = create_excel_report(export_df)
        pdf = create_pdf_report(export_df, st.session_state.motto)
        
        fn = f"Rozpis_{export_df.columns[1]}_{export_df.columns[-1]}"
        c1, c2 = st.columns(2)
        c1.download_button("⬇️ XLSX", xlsx, f"{fn}.xlsx")
        c2.download_button("⬇️ PDF", pdf, f"{fn}.pdf", mime="application/pdf")
        
        # Samostatne odosielanie pre hlavny rozpis
        with st.expander("📧 Email - Hlavný rozpis"):
            to = st.text_input("Komu (Rozpis):", st.session_state.config['email_settings']['default_to'])
            if st.button("Odoslať Rozpis"):
                ok, err = send_email_with_pdf(pdf, f"{fn}.pdf", to, st.session_state.config['email_settings']['default_subject'], st.session_state.config['email_settings']['default_body'])
                if ok:
                    st.success(f"Rozpis odoslaný na {to}")
                else:
                     st.error(f"Chyba pri odosielaní rozpisu: {err}")

elif mode == "⚙️ Nastavenia lekárov":
    st.header("Lekári")
    history_cache = load_history()
    c1, c2 = st.columns([3, 1])
    n = c1.text_input("Meno:")
    if c2.button("Pridať") and n:
        st.session_state.config['lekari'][n] = {"moze": ["Oddelenie"], "active": True, "short_term_active": False, "extra_dni": [], "priority": 100}
        save_config(st.session_state.config)
        st.rerun()
    
    for d in list(st.session_state.config['lekari'].keys()):
        p = st.session_state.config['lekari'][d]
        with st.expander(d):
            a = st.checkbox("Aktívny", p.get('active', True), key=f"a_{d}")
            sta = st.checkbox("Kratkodobo aktivny", p.get('short_term_active', False), key=f"sta_{d}")
            m = st.multiselect("Môže:", list(st.session_state.config['ambulancie'].keys())+["Oddelenie"], p.get('moze', []), key=f"m_{d}")
            prio = st.number_input("Priorita lekára (nižšie = skôr)", min_value=1, max_value=999, value=int(p.get('priority', 100)), key=f"prio_{d}")
            extra_dni_current = sorted(set(p.get('extra_dni', [])))

            st.caption("Dátumy krátkodobej aktivity")
            if extra_dni_current:
                pretty_dates = [datetime.strptime(x, '%Y-%m-%d').strftime('%d.%m.%Y') for x in extra_dni_current]
                st.write(", ".join(pretty_dates))
            else:
                st.write("Žiadne dátumy")

            dr = st.date_input("Pridať rozsah dátumov:", value=[], key=f"sta_range_{d}")
            c_add_dates, c_clear_dates = st.columns(2)
            add_dates_clicked = c_add_dates.button("Pridať dátumy", key=f"add_dates_{d}")
            clear_dates_clicked = c_clear_dates.button("Vymazať dátumy", key=f"clear_dates_{d}")

            if add_dates_clicked and dr:
                start_dt = dr[0]
                end_dt = dr[1] if len(dr) > 1 else dr[0]
                if end_dt < start_dt:
                    start_dt, end_dt = end_dt, start_dt
                new_dates = set(extra_dni_current)
                cur = start_dt
                while cur <= end_dt:
                    new_dates.add(cur.strftime('%Y-%m-%d'))
                    cur += timedelta(days=1)
                p['extra_dni'] = sorted(new_dates)
                save_config(st.session_state.config)
                st.rerun()

            if clear_dates_clicked and extra_dni_current:
                p['extra_dni'] = []
                save_config(st.session_state.config)
                st.rerun()

            c_rename, c_remove = st.columns(2)
            new_name = c_rename.text_input("Premenovať na:", value=d, key=f"rename_{d}")
            rename_clicked = c_rename.button("Premenovať", key=f"btn_rename_{d}")
            remove_clicked = c_remove.button("Odstrániť lekára", key=f"btn_remove_{d}")

            if rename_clicked:
                ok, msg = rename_doctor_everywhere(st.session_state.config, history_cache, st.session_state.manual_core, d, new_name)
                if ok:
                    save_config(st.session_state.config)
                    save_history(history_cache)
                    st.rerun()
                else:
                    st.error(msg)

            if remove_clicked:
                ok, msg = remove_doctor_everywhere(st.session_state.config, history_cache, st.session_state.manual_core, d)
                if ok:
                    save_config(st.session_state.config)
                    save_history(history_cache)
                    st.rerun()
                else:
                    st.error(msg)

            if (
                a!=p.get('active', True)
                or sta!=p.get('short_term_active', False)
                or m!=p.get('moze', [])
                or int(prio)!=int(p.get('priority', 100))
            ):
                p['active'], p['short_term_active'], p['moze'], p['priority'] = a, sta, m, int(prio)
                save_config(st.session_state.config)

elif mode == "🏥 Nastavenia ambulancií":
    st.header("Ambulancie")
    sel = st.selectbox("Vyber:", list(st.session_state.config['ambulancie'].keys()))
    curr = st.session_state.config['ambulancie'][sel]
    if isinstance(curr['priority'], list):
        txt = st.text_area("Priority:", ", ".join(curr['priority']))
        if st.button("Uložiť"):
            curr['priority'] = [x.strip() for x in txt.split(',')]
            save_config(st.session_state.config)

elif mode == "📧 Nastavenia Emailu":
    st.header("📧 Nastavenia Emailu")
    
    st.subheader("1. Hlavný rozpis služieb")
    c = st.session_state.config['email_settings']
    t = st.text_input("Predvolený príjemca (Rozpis):", c.get('default_to', ''))
    s = st.text_input("Predmet (Rozpis):", c.get('default_subject', ''))
    b = st.text_area("Text (Rozpis):", c.get('default_body', ''))
    
    st.markdown("---")
    
    st.subheader("2. Prehľad neprítomností")
    c_abs = st.session_state.config.get('email_settings_absences', {})
    t_abs = st.text_input("Predvolený príjemca (Absencie):", c_abs.get('default_to', ''))
    s_abs = st.text_input("Predmet (Absencie):", c_abs.get('default_subject', ''))
    b_abs = st.text_area("Text (Absencie):", c_abs.get('default_body', ''))

    if st.button("💾 Uložiť všetky nastavenia emailu"):
        st.session_state.config['email_settings'] = {"default_to": t, "default_subject": s, "default_body": b}
        st.session_state.config['email_settings_absences'] = {"default_to": t_abs, "default_subject": s_abs, "default_body": b_abs}
        save_config(st.session_state.config)
        st.success("Nastavenia uložené")
